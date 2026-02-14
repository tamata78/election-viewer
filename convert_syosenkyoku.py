#!/usr/bin/env python3
import openpyxl
import json
import re
import os

os.chdir('/Users/tamata78/work/election-viewer/temp_excel')

def clean_name(name):
    """区市町村名をクリーンアップ"""
    if not name:
        return None
    name = str(name).strip()
    name = re.sub(r'^[\s★☆　]+', '', name)
    return name

def get_type(name):
    """区部/市部を判定"""
    if '区' in name and '選挙区' not in name:
        return '区部'
    elif '市' in name or '町' in name or '村' in name:
        return '市部'
    return None

def convert_syosenkyoku_2024():
    """2024年小選挙区データを変換"""
    # 得票率ファイル
    wb_rate = openpyxl.load_workbook('shou_2024_rate.xlsx')
    ws_rate = wb_rate.active

    # 当選人数ファイル
    wb_seats = openpyxl.load_workbook('shou_2024_seats.xlsx')
    ws_seats = wb_seats.active

    # 政党リスト（5行目から取得）
    parties = []
    party_cols = {}
    row5 = [ws_rate.cell(row=5, column=c).value for c in range(1, 26)]
    print(f"Row 5: {row5}")

    # 政党名を抽出（全党派計の後から）
    col = 4  # D列から
    while col <= 25:
        party = ws_rate.cell(row=5, column=col).value
        if party and party not in ['全党派計', None, '']:
            parties.append(party)
            party_cols[party] = col
            col += 2  # 票数と率の2列
        else:
            col += 1
            if col > 25:
                break

    print(f"政党: {parties}")
    print(f"政党列: {party_cols}")

    # 当選人数を取得
    seats = {}
    for row in range(7, ws_seats.max_row + 1):
        district = ws_seats.cell(row=row, column=1).value
        if not district:
            continue
        district = clean_name(district)
        if not district or '計' in district:
            continue

        # 各政党の当選人数（計の列を探す）
        for party in parties:
            # 当選人数ファイルの列構造を確認
            pass

    # 都計から政党別合計を取得
    total = {"totalVotes": 0}
    for row in range(7, ws_rate.max_row + 1):
        name = ws_rate.cell(row=row, column=1).value
        if name and '都計' in str(name):
            total_votes = ws_rate.cell(row=row, column=2).value
            total["totalVotes"] = int(float(total_votes)) if total_votes else 0

            for party in parties:
                col = party_cols[party]
                votes = ws_rate.cell(row=row, column=col).value
                rate = ws_rate.cell(row=row, column=col+1).value
                total[party] = {
                    "votes": int(float(votes)) if votes else 0,
                    "rate": round(float(rate), 2) if rate else 0,
                    "seats": 0  # 後で更新
                }
            break

    # 当選人数を追加
    # 当選人数ファイルを再解析
    print("\n=== 当選人数ファイルの構造 ===")
    for row in range(5, 10):
        row_data = [ws_seats.cell(row=row, column=c).value for c in range(1, 20)]
        print(f"Row {row}: {row_data}")

    # 当選人数の集計（7行目から、☆がある選挙区行のみ）
    total_seats = {}
    for party in parties:
        total_seats[party] = 0

    # 各政党の計の列（2024年用）
    party_seat_cols = {
        '自由民主党': 10,
        '立憲民主党': 14,
        '公明党': 18,
        '日本維新の会': 22,
        '日本共産党': 26,
        '参政党': 30,
        '国民民主党': 34,
        'みんなでつくる党': 38,
        'れいわ新選組': 42,
        '本人届出': 46,
    }

    for row in range(7, ws_seats.max_row + 1):
        district = ws_seats.cell(row=row, column=1).value
        # ☆がある行のみカウント（合計行を除外）
        if not district or '☆' not in str(district):
            continue

        for party, col in party_seat_cols.items():
            if party in total:
                val = ws_seats.cell(row=row, column=col).value
                if val and isinstance(val, (int, float)) and val > 0:
                    total_seats[party] += int(val)

    # 合計を更新
    for party in parties:
        if party in total and party in total_seats:
            total[party]["seats"] = total_seats[party]

    print(f"\n当選人数: {total_seats}")

    # 区市町村別データ
    municipalities = []
    current_district = ""

    for row in range(7, ws_rate.max_row + 1):
        name_raw = ws_rate.cell(row=row, column=1).value
        name = clean_name(name_raw)

        if not name:
            continue

        # 選挙区行を検出
        if '区' in name and (name.startswith('☆') or re.match(r'^\d+区', name) or '☆' in str(name_raw)):
            match = re.search(r'(\d+)区', name)
            if match:
                current_district = f"{match.group(1)}区"
            continue

        # 計の行はスキップ
        if '計' in name:
            continue

        region_type = get_type(name)
        if not region_type:
            continue

        total_votes = ws_rate.cell(row=row, column=2).value
        muni = {
            "name": name,
            "district": current_district,
            "type": region_type,
            "totalVotes": int(float(total_votes)) if total_votes else 0
        }

        for party in parties:
            col = party_cols[party]
            votes = ws_rate.cell(row=row, column=col).value
            rate = ws_rate.cell(row=row, column=col+1).value
            muni[party] = {
                "votes": int(float(votes)) if votes else 0,
                "rate": round(float(rate), 2) if rate else 0
            }

        municipalities.append(muni)

    result = {
        "electionType": "小選挙区",
        "electionDate": "2024-10-27",
        "parties": parties,
        "total": total,
        "municipalities": municipalities
    }

    return result

def convert_syosenkyoku_2026():
    """2026年小選挙区データを変換"""
    wb_rate = openpyxl.load_workbook('shou_2026_rate.xlsx')
    ws_rate = wb_rate.active

    wb_seats = openpyxl.load_workbook('shou_2026_seats_new.xlsx')
    ws_seats = wb_seats.active

    print("\n=== 2026年小選挙区 ===")

    # ヘッダー確認
    for row in range(1, 10):
        row_data = [ws_rate.cell(row=row, column=c).value for c in range(1, 20)]
        print(f"Row {row}: {row_data}")

    # 政党リスト
    parties = []
    party_cols = {}
    col = 4
    while col <= 25:
        party = ws_rate.cell(row=5, column=col).value
        if party and party not in ['全党派計', None, '']:
            parties.append(party)
            party_cols[party] = col
            col += 2
        else:
            col += 1
            if col > 25:
                break

    print(f"2026政党: {parties}")

    # 当選人数を集計（seats fileから）
    party_seat_cols = {
        '自由民主党': 10,
        '参政党': 14,
        '国民民主党': 18,
        '中道改革連合': 22,
        '日本共産党': 26,
        '日本維新の会': 30,
        'チームみらい': 34,
        'れいわ新選組': 38,
        '日本保守党': 42,
        '減税日本・ゆうこく連合': 46,
        '本人届出': 50,
    }
    total_seats = {party: 0 for party in party_seat_cols}

    for row in range(7, ws_seats.max_row + 1):
        district = ws_seats.cell(row=row, column=1).value
        if district and '☆' in str(district):
            for party, col in party_seat_cols.items():
                val = ws_seats.cell(row=row, column=col).value
                if val and isinstance(val, (int, float)) and val > 0:
                    total_seats[party] += int(val)
                elif val and str(val).isdigit() and int(val) > 0:
                    total_seats[party] += int(val)

    print(f"2026当選人数: {total_seats}")

    # 都計から合計取得
    total = {"totalVotes": 0}
    for row in range(7, ws_rate.max_row + 1):
        name = ws_rate.cell(row=row, column=1).value
        if name and '都計' in str(name):
            total_votes = ws_rate.cell(row=row, column=2).value
            total["totalVotes"] = int(float(total_votes)) if total_votes else 0

            for party in parties:
                col = party_cols[party]
                votes = ws_rate.cell(row=row, column=col).value
                rate = ws_rate.cell(row=row, column=col+1).value
                seats = total_seats.get(party, 0)
                total[party] = {
                    "votes": int(float(votes)) if votes else 0,
                    "rate": round(float(rate), 2) if rate else 0,
                    "seats": seats
                }
            break

    # 区市町村別データ
    municipalities = []
    current_district = ""

    for row in range(7, ws_rate.max_row + 1):
        name_raw = ws_rate.cell(row=row, column=1).value
        name = clean_name(name_raw)

        if not name:
            continue

        if '区' in name and (name.startswith('☆') or re.match(r'^\d+区', name) or '☆' in str(name_raw)):
            match = re.search(r'(\d+)区', name)
            if match:
                current_district = f"{match.group(1)}区"
            continue

        if '計' in name:
            continue

        region_type = get_type(name)
        if not region_type:
            continue

        total_votes = ws_rate.cell(row=row, column=2).value
        muni = {
            "name": name,
            "district": current_district,
            "type": region_type,
            "totalVotes": int(float(total_votes)) if total_votes else 0
        }

        for party in parties:
            col = party_cols[party]
            votes = ws_rate.cell(row=row, column=col).value
            rate = ws_rate.cell(row=row, column=col+1).value
            muni[party] = {
                "votes": int(float(votes)) if votes else 0,
                "rate": round(float(rate), 2) if rate else 0
            }

        municipalities.append(muni)

    result = {
        "electionType": "小選挙区",
        "electionDate": "2026-02-08",
        "parties": parties,
        "total": total,
        "municipalities": municipalities
    }

    return result

if __name__ == '__main__':
    # 2024年
    print("=== 2024年小選挙区 ===")
    shou_2024 = convert_syosenkyoku_2024()
    with open('../public/data/tokyo-syosenkyoku-2024.json', 'w', encoding='utf-8') as f:
        json.dump(shou_2024, f, ensure_ascii=False, indent=2)
    print(f"Saved: tokyo-syosenkyoku-2024.json ({len(shou_2024['municipalities'])} municipalities)")

    # 2026年
    print("\n=== 2026年小選挙区 ===")
    shou_2026 = convert_syosenkyoku_2026()
    with open('../public/data/tokyo-syosenkyoku-2026.json', 'w', encoding='utf-8') as f:
        json.dump(shou_2026, f, ensure_ascii=False, indent=2)
    print(f"Saved: tokyo-syosenkyoku-2026.json ({len(shou_2026['municipalities'])} municipalities)")
