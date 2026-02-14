#!/usr/bin/env python3
import openpyxl
import json
import re

def clean_name(name):
    """区市町村名をクリーンアップ"""
    if not name:
        return None
    name = str(name).strip()
    # 先頭のスペースや記号を除去
    name = re.sub(r'^[\s★☆　]+', '', name)
    return name

def get_type(name):
    """区部/市部を判定"""
    if '区' in name and '選挙区' not in name:
        return '区部'
    elif '市' in name or '町' in name or '村' in name:
        return '市部'
    return None

def convert_hirei_2024():
    """2024年比例代表データを変換"""
    wb = openpyxl.load_workbook('hirei_2024_votes.xlsx')
    ws = wb.active

    # 政党名（6行目）
    parties = []
    party_cols = {}
    for col in range(3, 14):  # C列からM列
        party_name = ws.cell(row=6, column=col).value
        if party_name:
            parties.append(party_name)
            party_cols[party_name] = col

    print(f"政党: {parties}")

    # データ抽出
    municipalities = []
    total = {}

    for row in range(9, ws.max_row + 1):
        name_raw = ws.cell(row=row, column=1).value
        name = clean_name(name_raw)

        if not name:
            continue

        # 合計行
        if '都計' in name:
            for party in parties:
                col = party_cols[party]
                val = ws.cell(row=row, column=col).value
                total[party] = int(float(val)) if val else 0
            total['合計'] = int(float(ws.cell(row=row, column=14).value))
            continue

        # 区部計・市部計はスキップ
        if '計' in name:
            continue

        region_type = get_type(name)
        if not region_type:
            continue

        votes = {}
        row_total = 0
        for party in parties:
            col = party_cols[party]
            val = ws.cell(row=row, column=col).value
            votes[party] = int(float(val)) if val else 0
            row_total += votes[party]

        # 合計列から取得
        total_val = ws.cell(row=row, column=14).value
        if total_val:
            row_total = int(float(total_val))

        municipalities.append({
            "name": name,
            "type": region_type,
            "votes": votes,
            "total": row_total
        })

    result = {
        "electionType": "比例代表",
        "electionDate": "2024-10-27",
        "parties": [{"id": i+1, "name": p} for i, p in enumerate(parties)],
        "total": total,
        "municipalities": municipalities
    }

    return result

def convert_syosenkyoku_2024():
    """2024年小選挙区データを変換"""
    # 得票率ファイルから読み込み
    wb = openpyxl.load_workbook('shou_2024_rate.xlsx')
    ws = wb.active

    print(f"小選挙区: Sheet names: {wb.sheetnames}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

    # 最初の20行を確認
    for row in range(1, 20):
        row_data = [ws.cell(row=row, column=c).value for c in range(1, 15)]
        print(f"Row {row}: {row_data}")

    return None

def convert_syosenkyoku_breakdown_2024():
    """2024年小選挙区 開票結果内訳を変換"""
    wb = openpyxl.load_workbook('shou_2024_breakdown.xlsx')
    ws = wb.active

    print(f"小選挙区内訳: Max row: {ws.max_row}, Max col: {ws.max_column}")

    # 最初の25行を確認
    for row in range(1, 25):
        row_data = [ws.cell(row=row, column=c).value for c in range(1, 20)]
        print(f"Row {row}: {row_data}")

    return None

def convert_syosenkyoku_seats_2024():
    """2024年小選挙区 当選人数を変換"""
    wb = openpyxl.load_workbook('shou_2024_seats.xlsx')
    ws = wb.active

    print(f"当選人数: Max row: {ws.max_row}, Max col: {ws.max_column}")

    # 最初の15行を確認
    for row in range(1, 15):
        row_data = [ws.cell(row=row, column=c).value for c in range(1, 15)]
        print(f"Row {row}: {row_data}")

    return None

if __name__ == '__main__':
    import os
    os.chdir('/Users/tamata78/work/election-viewer/temp_excel')

    print("=== 2024年比例代表 ===")
    hirei_2024 = convert_hirei_2024()

    with open('../public/data/tokyo-hirei-2024.json', 'w', encoding='utf-8') as f:
        json.dump(hirei_2024, f, ensure_ascii=False, indent=2)
    print(f"Saved: tokyo-hirei-2024.json ({len(hirei_2024['municipalities'])} municipalities)")

    print("\n=== 2024年小選挙区（得票率）===")
    convert_syosenkyoku_2024()

    print("\n=== 2024年小選挙区（内訳）===")
    convert_syosenkyoku_breakdown_2024()

    print("\n=== 2024年小選挙区（当選人数）===")
    convert_syosenkyoku_seats_2024()
