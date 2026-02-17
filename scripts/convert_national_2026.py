#!/usr/bin/env python3
"""
2026年衆議院選挙 全国データ変換スクリプト
Excel → NationalElectionData JSON
"""
import json
import re
import openpyxl

SHOU_FILE = '/Users/tamata78/Downloads/2026_都道府県別党派別新前元別当選人数（小選挙区）.xlsx'
HIREI_FILE = '/Users/tamata78/Downloads/2026_比例代表当選人数.xlsx'
OUTPUT = '/Users/tamata78/work/election-viewer/public/data/elections/shugiin_2026.json'

PREFECTURES = [
    '北海道', '青森県', '岩手県', '宮城県', '秋田県', '山形県', '福島県',
    '茨城県', '栃木県', '群馬県', '埼玉県', '千葉県', '東京都', '神奈川県',
    '新潟県', '富山県', '石川県', '福井県', '山梨県', '長野県',
    '岐阜県', '静岡県', '愛知県', '三重県',
    '滋賀県', '京都府', '大阪府', '兵庫県', '奈良県', '和歌山県',
    '鳥取県', '島根県', '岡山県', '広島県', '山口県',
    '徳島県', '香川県', '愛媛県', '高知県',
    '福岡県', '佐賀県', '長崎県', '熊本県', '大分県', '宮崎県', '鹿児島県', '沖縄県'
]

BLOCKS = {
    '北海道': ['北海道'],
    '東北': ['青森県', '岩手県', '宮城県', '秋田県', '山形県', '福島県'],
    '北関東': ['茨城県', '栃木県', '群馬県', '埼玉県'],
    '南関東': ['千葉県', '神奈川県', '山梨県'],
    '東京都': ['東京都'],
    '北陸信越': ['新潟県', '富山県', '石川県', '福井県', '長野県'],
    '東海': ['岐阜県', '静岡県', '愛知県', '三重県'],
    '近畿': ['滋賀県', '京都府', '大阪府', '兵庫県', '奈良県', '和歌山県'],
    '中国': ['鳥取県', '島根県', '岡山県', '広島県', '山口県'],
    '四国': ['徳島県', '香川県', '愛媛県', '高知県'],
    '九州': ['福岡県', '佐賀県', '長崎県', '熊本県', '大分県', '宮崎県', '鹿児島県', '沖縄県'],
}

BLOCK_SEATS = {
    '北海道': 8, '東北': 12, '北関東': 19, '南関東': 23,
    '東京都': 19, '北陸信越': 10, '東海': 21, '近畿': 28,
    '中国': 10, '四国': 6, '九州': 20,
}


def safe_int(val):
    """Safely convert a value to int."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace(',', '').replace(' ', '').replace('\u3000', '').replace('票', '')
    # Take first number from string
    m = re.search(r'\d+', s)
    return int(m.group(0)) if m else 0


def parse_count(val):
    """Parse elected count from value like '1                        1' or integer."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    nums = re.findall(r'\d+', s)
    if nums:
        return int(nums[-1])  # 計 is the last number
    return 0


def detect_block(r0, r1, r2):
    """Detect block name from first 3 columns."""
    block_id = (str(r0 or '') + str(r1 or '') + str(r2 or '')).replace(' ', '').replace('\u3000', '')
    if not block_id or block_id == 'NoneNoneNone':
        return None
    mapping = [
        ('北海道', '北海道'), ('東北', '東北'), ('北関東', '北関東'),
        ('南関東', '南関東'), ('東京', '東京都'), ('北陸', '北陸信越'),
        ('東海', '東海'), ('近畿', '近畿'), ('中国', '中国'),
        ('四国', '四国'), ('九州', '九州'),
    ]
    for key, name in mapping:
        if key in block_id:
            return name
    return None


def extract_shou_data():
    """Extract 小選挙区 data from the Excel file."""
    wb = openpyxl.load_workbook(SHOU_FILE)

    # --- Table 17/18: elected counts per prefecture ---
    rows17 = list(wb['Table 17'].iter_rows(min_row=3, max_row=49, values_only=True))
    rows18 = list(wb['Table 18'].iter_rows(min_row=3, max_row=49, values_only=True))

    # --- Table 24/25: votes per prefecture (main parties, clean format) ---
    # Table 24: pref(3), LDP(男女計), chudo(男女計), ishin(男女計), kokumin(男女計)
    rows24 = list(wb['Table 24'].iter_rows(min_row=3, max_row=49, values_only=True))
    # Table 25: sansei(男女計), kyosan(男女計), reiwa(男女計), genzei(男女計)
    rows25 = list(wb['Table 25'].iter_rows(min_row=3, max_row=49, values_only=True))

    # --- Table 31: total valid votes per prefecture ---
    rows31 = list(wb['Table 31'].iter_rows(min_row=2, max_row=48, values_only=True))

    prefectures = []
    for i, pref_name in enumerate(PREFECTURES):
        r17 = rows17[i]
        total_districts = safe_int(r17[3])

        # Elected counts from Table 17
        # Note: merged cells cause 維新計 to appear at col 12 (not 13)
        elected = {
            '自由民主党': parse_count(r17[7]),
            '中道改革連合': parse_count(r17[8]),
            '日本維新の会': parse_count(r17[12]) or parse_count(r17[13]),
            '国民民主党': parse_count(r17[14]),
            '参政党': parse_count(r17[15]),
            '日本共産党': parse_count(r17[16]),
            'れいわ新選組': parse_count(r17[17]),
        }

        # Elected counts from Table 18
        # 無所属 has 新/前/元/計 across cols 8-13, but merged cells shift 計
        r18 = rows18[i]
        elected['減税日本・ゆうこく連合'] = parse_count(r18[4])
        elected['日本保守党'] = parse_count(r18[5])
        elected['社会民主党'] = parse_count(r18[6])
        elected['チームみらい'] = parse_count(r18[7])
        # 無所属: take last non-zero value in cols 8-13 range as 計
        musozoku = 0
        for ci in range(13, 7, -1):
            if ci < len(r18) and r18[ci] is not None:
                musozoku = parse_count(r18[ci])
                break
        elected['無所属'] = musozoku

        # Votes from Table 24 (計 columns: 5, 8, 11, 14)
        r24 = rows24[i]
        votes = {
            '自由民主党': safe_int(r24[5]),
            '中道改革連合': safe_int(r24[8]),
            '日本維新の会': safe_int(r24[11]),
            '国民民主党': safe_int(r24[14]),
        }

        # Votes from Table 25 (計 columns: 5, 8, 11, 14)
        r25 = rows25[i]
        votes['参政党'] = safe_int(r25[5])
        votes['日本共産党'] = safe_int(r25[8])
        votes['れいわ新選組'] = safe_int(r25[11])
        votes['減税日本・ゆうこく連合'] = safe_int(r25[14])

        # Total valid votes from Table 31
        r31 = rows31[i]
        total_votes_pref = safe_int(r31[2])  # 有効投票数
        if total_votes_pref == 0:
            total_votes_pref = sum(votes.values())

        # Build party results (skip parties with 0 votes and 0 seats)
        all_parties = [
            '自由民主党', '中道改革連合', '日本維新の会', '国民民主党',
            '参政党', '日本共産党', 'れいわ新選組', '減税日本・ゆうこく連合',
            '日本保守党', '社会民主党', 'チームみらい', '無所属',
        ]
        party_results = []
        for party in all_parties:
            v = votes.get(party, 0)
            e = elected.get(party, 0)
            if v > 0 or e > 0:
                rate = (v / total_votes_pref * 100) if total_votes_pref > 0 else 0
                party_results.append({
                    'party': party,
                    'seats': e,
                    'totalVotes': v,
                    'voteRate': round(rate, 2),
                })

        party_results.sort(key=lambda x: (-x['seats'], -x['totalVotes']))

        prefectures.append({
            'prefecture': pref_name,
            'totalDistricts': total_districts,
            'partyResults': party_results,
        })

    return prefectures


def extract_hirei_data():
    """Extract 比例代表 data."""
    wb_shou = openpyxl.load_workbook(SHOU_FILE)

    # --- Block-level votes from Table 28/29/30 (比例代表 by block) ---
    block_votes = {}
    block_totals = {}

    def parse_block_table(ws, party_names, target_dict, col_offset=6):
        """Parse a block-structured table to extract per-block party data."""
        current_block = None
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            new_block = detect_block(row[0], row[1], row[2])
            if new_block:
                current_block = new_block
            if not current_block:
                continue

            r3 = str(row[3] or '').replace(' ', '').replace('\u3000', '')
            is_total = '計' in r3
            is_single_block = current_block in ('北海道', '東京都') and r3 and not is_total and '選挙' not in r3 and '都道' not in r3

            if is_total or is_single_block:
                if current_block not in target_dict:
                    target_dict[current_block] = {}
                for j, party in enumerate(party_names):
                    col = col_offset + j
                    if col < len(row):
                        target_dict[current_block][party] = safe_int(row[col])

    # Table 28: LDP, chudo, ishin, kokumin, sansei, kyosan (cols 6-11)
    parse_block_table(wb_shou['Table 28'],
                      ['自由民主党', '中道改革連合', '日本維新の会', '国民民主党', '参政党', '日本共産党'],
                      block_votes)

    # Table 29: reiwa, genzei, hosyu, sdp, team_mirai (cols 6-10)
    extra_votes = {}
    parse_block_table(wb_shou['Table 29'],
                      ['れいわ新選組', '減税日本・ゆうこく連合', '日本保守党', '社会民主党', 'チームみらい'],
                      extra_votes)
    for block in extra_votes:
        if block in block_votes:
            block_votes[block].update(extra_votes[block])
        else:
            block_votes[block] = extra_votes[block]

    # Table 30: total votes (col 6)
    current_block = None
    for row in wb_shou['Table 30'].iter_rows(min_row=2, max_row=58, values_only=True):
        new_block = detect_block(row[0], row[1], row[2])
        if new_block:
            current_block = new_block
        if not current_block:
            continue
        r3 = str(row[3] or '').replace(' ', '').replace('\u3000', '')
        is_total = '計' in r3
        is_single = current_block in ('北海道', '東京都') and r3 and not is_total and '選挙' not in r3 and '都道' not in r3
        if is_total or is_single:
            block_totals[current_block] = safe_int(row[6])

    # --- Parse hirei Excel for elected counts per party per block ---
    wb_hirei = openpyxl.load_workbook(HIREI_FILE)

    # Each block has 2-4 tables containing party data side-by-side
    block_table_map = {
        '北海道': [1, 2, 3],
        '東北': [4, 5, 6],
        '北関東': [7, 8, 9],
        '南関東': [10, 11, 12],
        '東京都': [13, 14, 15],
        '北陸信越': [16, 17, 18],
        '東海': [19, 20, 21],
        '近畿': [22, 23, 24],
        '中国': [25, 26, 27],
        '四国': [28, 29],
        '九州': [30, 31, 32, 33],
    }

    known_parties = [
        '自由民主党', '中道改革連合', '日本維新の会', '国民民主党',
        '参政党', '日本共産党', 'れいわ新選組', '社会民主党',
        'チームみらい', '日本保守党', '減税日本・ゆうこく連合',
    ]

    def parse_hirei_tables(table_nums):
        """Parse hirei tables for a block, extracting party elected counts and votes."""
        parties = {}
        for tnum in table_nums:
            ws = wb_hirei[f'Table {tnum}']
            all_rows = list(ws.iter_rows(min_row=1, max_row=min(3, ws.max_row),
                                         max_col=ws.max_column, values_only=False))
            if len(all_rows) < 3:
                continue

            # Identify parties from row 1
            row1 = [(c.value, c.column) for c in all_rows[0]]
            row2 = [(c.value, c.column) for c in all_rows[1]]
            row3 = [(c.value, c.column) for c in all_rows[2]]

            party_cols = []
            for val, col in row1:
                if val:
                    for p in known_parties:
                        if p in str(val):
                            party_cols.append((p, col))
                            break

            for party_name, start_col in party_cols:
                # Elected count from row 3
                seats = 0
                for val, col in row3:
                    if val and col >= start_col and col < start_col + 10:
                        m = re.search(r'(\d+)\s*人', str(val))
                        if m:
                            seats = int(m.group(1))
                            break

                # Vote count from row 2
                vote_count = 0
                for val, col in row2:
                    if val and col >= start_col and col < start_col + 10:
                        if isinstance(val, (int, float)):
                            vote_count = int(val)
                            break
                        s = str(val).replace(',', '').replace(' ', '').replace('票', '')
                        m = re.search(r'(\d+)', s)
                        if m:
                            vote_count = int(m.group(1))
                            break

                if party_name not in parties:
                    parties[party_name] = {'seats': seats, 'votes': vote_count}
                else:
                    if seats > 0:
                        parties[party_name]['seats'] = seats
                    if vote_count > 0:
                        parties[party_name]['votes'] = vote_count

        return parties

    # Build hirei blocks
    blocks = []
    for block_name in BLOCKS:
        tables = block_table_map.get(block_name, [])
        party_data = parse_hirei_tables(tables)

        total_votes = block_totals.get(block_name, 0)

        # Merge block_votes (from shou Excel) for any missing vote data
        bv = block_votes.get(block_name, {})
        for party_name, v in bv.items():
            if v > 0:
                if party_name in party_data:
                    if party_data[party_name]['votes'] == 0:
                        party_data[party_name]['votes'] = v
                else:
                    party_data[party_name] = {'seats': 0, 'votes': v}

        if total_votes == 0:
            total_votes = sum(d['votes'] for d in party_data.values())

        hirei_parties = []
        for party_name, data in party_data.items():
            v = data['votes']
            rate = (v / total_votes * 100) if total_votes > 0 else 0
            hirei_parties.append({
                'party': party_name,
                'block': block_name,
                'seats': data['seats'],
                'votes': v,
                'voteRate': round(rate, 2),
                'candidates': [],
            })

        hirei_parties.sort(key=lambda x: (-x['seats'], -x['votes']))

        blocks.append({
            'name': block_name,
            'totalSeats': BLOCK_SEATS.get(block_name, 0),
            'totalVotes': total_votes,
            'parties': hirei_parties,
        })

    return blocks


def main():
    print('Extracting 小選挙区 data...')
    prefectures = extract_shou_data()

    print('Extracting 比例代表 data...')
    hirei_blocks = extract_hirei_data()

    total_shou_seats = sum(p['totalDistricts'] for p in prefectures)
    total_hirei_seats = sum(b['totalSeats'] for b in hirei_blocks)

    data = {
        'year': 2026,
        'electionDate': '2026-02-08',
        'hirei': {
            'totalSeats': total_hirei_seats,
            'blocks': hirei_blocks,
        },
        'shou': {
            'totalSeats': total_shou_seats,
            'prefectures': prefectures,
            'districts': [],
        },
    }

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f'\nOutput: {OUTPUT}')
    print(f'小選挙区: {total_shou_seats} seats, {len(prefectures)} prefectures')
    print(f'比例代表: {total_hirei_seats} seats, {len(hirei_blocks)} blocks')

    # Validation
    print('\n--- Validation ---')
    shou_ok = True
    for p in prefectures:
        total_elected = sum(r['seats'] for r in p['partyResults'])
        if total_elected != p['totalDistricts']:
            print(f"  WARN {p['prefecture']}: elected={total_elected} != districts={p['totalDistricts']}")
            shou_ok = False
    if shou_ok:
        print('  小選挙区: All prefectures OK')

    hirei_ok = True
    for b in hirei_blocks:
        total_elected = sum(p['seats'] for p in b['parties'])
        if total_elected != b['totalSeats']:
            print(f"  WARN {b['name']}: elected={total_elected} != seats={b['totalSeats']}")
            hirei_ok = False
    if hirei_ok:
        print('  比例代表: All blocks OK')

    # Summary
    print('\n--- 小選挙区 当選者数 ---')
    party_totals = {}
    for p in prefectures:
        for r in p['partyResults']:
            party_totals[r['party']] = party_totals.get(r['party'], 0) + r['seats']
    for party, seats in sorted(party_totals.items(), key=lambda x: -x[1]):
        if seats > 0:
            print(f'  {party}: {seats}')

    print('\n--- 比例代表 当選者数 ---')
    hirei_totals = {}
    for b in hirei_blocks:
        for p in b['parties']:
            hirei_totals[p['party']] = hirei_totals.get(p['party'], 0) + p['seats']
    for party, seats in sorted(hirei_totals.items(), key=lambda x: -x[1]):
        if seats > 0:
            print(f'  {party}: {seats}')


if __name__ == '__main__':
    main()
