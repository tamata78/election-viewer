#!/usr/bin/env python3
"""
2024年衆議院選挙 全国データ変換スクリプト
Excel → NationalElectionData JSON
"""
import json
import re
import openpyxl

EXCEL_FILE = '/Users/tamata78/Downloads/2024_衆議員選_小選挙区_比例区.xlsx'
OUTPUT = '/Users/tamata78/work/election-viewer/public/data/elections/shugiin_2024.json'

PREFECTURES = [
    '北海道', '青森県', '岩手県', '宮城県', '秋田県', '山形県', '福島県',
    '茨城県', '栃木県', '群馬県', '埼玉県', '千葉県', '東京都', '神奈川県',
    '新潟県', '富山県', '石川県', '福井県', '山梨県', '長野県',
    '岐阜県', '静岡県', '愛知県', '三重県',
    '滋賀県', '京都府', '大阪府', '兵庫県', '奈良県', '和歌山県',
    '鳥取県', '島根県', '岡山県', '広島県', '山口県',
    '徳島県', '香川県', '愛媛県', '高知県',
    '福岡県', '佐賀県', '長崎県', '熊本県', '大分県', '宮崎県', '鹿児島県', '沖縄県'
]

BLOCK_ORDER = ['北海道', '東北', '北関東', '南関東', '東京都', '北陸信越', '東海', '近畿', '中国', '四国', '九州']

# 2024 block seats (derived from actual elected counts in Excel)
BLOCK_SEATS = {
    '北海道': 8, '東北': 12, '北関東': 19, '南関東': 23,
    '東京都': 19, '北陸信越': 10, '東海': 21, '近畿': 28,
    '中国': 10, '四国': 6, '九州': 20,
}

# Hirei party tables → block mapping
HIREI_TABLE_MAP = {
    '北海道': list(range(41, 44)),
    '東北': list(range(44, 47)),
    '北関東': list(range(47, 50)),
    '南関東': list(range(50, 53)),
    '東京都': list(range(53, 56)),
    '北陸信越': list(range(56, 59)),
    '東海': list(range(59, 62)),
    '近畿': list(range(62, 65)),
    '中国': list(range(65, 68)),
    '四国': list(range(68, 70)),
    '九州': list(range(70, 74)),
}

# Hirei block ranking tables
HIREI_RANK_TABLE = {
    '北海道': 30, '東北': 31, '北関東': 32, '南関東': 33,
    '東京都': 34, '北陸信越': 35, '東海': 36, '近畿': 37,
    '中国': 38, '四国': 39, '九州': 40,
}

KNOWN_PARTIES = [
    '自由民主党', '立憲民主党', '日本維新の会', '公明党', '日本共産党', '国民民主党',
    'れいわ新選組', '社会民主党', '参政党', 'みんなでつくる党', '安楽死制度を考える会', '日本保守党',
]


def safe_int(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace(',', '').replace(' ', '').replace('\u3000', '').replace('票', '')
    m = re.search(r'\d+', s)
    return int(m.group(0)) if m else 0


def parse_count(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    nums = re.findall(r'\d+', s)
    if nums:
        return int(nums[-1])
    return 0


def last_nonzero_in_range(row, start, end):
    """Get last non-None numeric value in column range (inclusive)."""
    for col in range(end, start - 1, -1):
        if col < len(row) and row[col] is not None:
            v = parse_count(row[col])
            if v > 0:
                return v
    return 0


def extract_shou_data(wb):
    """Extract 小選挙区 data."""
    # Table 14: elected counts for major parties
    # Header row 1 cols: LDP=5, CDP=10, 維新=15, 公明=19, 共産=22, 国民=25, れいわ=31, 社民=32
    # 計 is the last non-None value in each party's column range
    rows14 = list(wb['Table 14'].iter_rows(min_row=3, max_row=49, values_only=True))

    # Table 16: independents
    rows16 = list(wb['Table 16'].iter_rows(min_row=4, max_row=50, values_only=True))

    # Table 21-24: votes
    rows21 = list(wb['Table 21'].iter_rows(min_row=3, max_row=49, values_only=True))
    rows22 = list(wb['Table 22'].iter_rows(min_row=3, max_row=49, values_only=True))
    rows23 = list(wb['Table 23'].iter_rows(min_row=3, max_row=49, values_only=True))
    rows24 = list(wb['Table 24'].iter_rows(min_row=3, max_row=49, values_only=True))

    # Table 28: total valid votes
    rows28 = list(wb['Table 28'].iter_rows(min_row=2, max_row=48, values_only=True))

    prefectures = []
    for i, pref_name in enumerate(PREFECTURES):
        r14 = rows14[i]
        total_districts = safe_int(r14[3])

        # Elected counts - use last_nonzero for each party range
        # Column numbers are 1-based, tuple indices are 0-based (index = col - 1)
        # LDP: cols 5-9 (indices 4-8), CDP: cols 10-14 (indices 9-13), etc.
        elected = {
            '自由民主党': last_nonzero_in_range(r14, 4, 8),
            '立憲民主党': last_nonzero_in_range(r14, 9, 13),
            '日本維新の会': last_nonzero_in_range(r14, 14, 17),
            '公明党': last_nonzero_in_range(r14, 18, 20),
            '日本共産党': last_nonzero_in_range(r14, 21, 23),
            '国民民主党': last_nonzero_in_range(r14, 24, 29),
            'れいわ新選組': parse_count(r14[30]) if len(r14) > 30 else 0,
            '社会民主党': last_nonzero_in_range(r14, 31, 35) if len(r14) > 31 else 0,
        }

        # Table 16: independents (無所属) and party-endorsed independents (推薦)
        r16 = rows16[i]
        # 推薦 計 in col 9 (index 8), 無所属 計 in cols 10-14 (indices 9-13)
        suisen = last_nonzero_in_range(r16, 4, 8)
        musozoku = last_nonzero_in_range(r16, 9, 13)
        elected['無所属'] = suisen + musozoku

        # Votes from Table 21 (LDP, CDP, 維新, 公明) - 計 cols: 5, 8, 11, 14
        r21 = rows21[i]
        votes = {
            '自由民主党': safe_int(r21[5]),
            '立憲民主党': safe_int(r21[8]),
            '日本維新の会': safe_int(r21[11]),
            '公明党': safe_int(r21[14]),
        }

        # Votes from Table 22 (共産, 国民, れいわ, 社民) - 計 cols: 5, 8, 11, 14
        r22 = rows22[i]
        votes['日本共産党'] = safe_int(r22[5])
        votes['国民民主党'] = safe_int(r22[8])
        votes['れいわ新選組'] = safe_int(r22[11])
        votes['社会民主党'] = safe_int(r22[14])

        # Votes from Table 23 (参政, みんなでつくる) - 計 cols: 5, 8
        r23 = rows23[i]
        votes['参政党'] = safe_int(r23[5])

        # Votes from Table 24 (諸派, 無所属, 合計) - 計 cols: 5, 8, 11
        r24 = rows24[i]
        votes['無所属'] = safe_int(r24[8])

        # Total valid votes from Table 28
        r28 = rows28[i]
        total_votes_pref = safe_int(r28[5])  # col varies; try 5 (有効投票数)
        if total_votes_pref == 0:
            total_votes_pref = safe_int(r28[4])
        if total_votes_pref == 0:
            total_votes_pref = sum(votes.values())

        # Build party results
        all_parties = [
            '自由民主党', '立憲民主党', '日本維新の会', '公明党', '日本共産党', '国民民主党',
            'れいわ新選組', '社会民主党', '参政党', '無所属',
        ]
        party_results = []
        for party in all_parties:
            v = votes.get(party, 0)
            e = elected.get(party, 0)
            if v > 0 or e > 0:
                rate = (v / total_votes_pref * 100) if total_votes_pref > 0 else 0
                party_results.append({
                    'party': party,
                    'seats': e,
                    'totalVotes': v,
                    'voteRate': round(rate, 2),
                })

        party_results.sort(key=lambda x: (-x['seats'], -x['totalVotes']))

        prefectures.append({
            'prefecture': pref_name,
            'totalDistricts': total_districts,
            'partyResults': party_results,
        })

    return prefectures


def extract_hirei_data(wb):
    """Extract 比例代表 data."""
    blocks = []

    for block_name in BLOCK_ORDER:
        # Get party votes from ranking table (Tables 30-40)
        rank_table = HIREI_RANK_TABLE[block_name]
        ws_rank = wb[f'Table {rank_table}']
        party_votes = {}
        total_votes = 0

        for row in ws_rank.iter_rows(min_row=3, values_only=True):
            if row[1] and row[1] == '得票総数':
                total_votes = safe_int(row[2])
            elif row[0] is not None and isinstance(row[0], (int, float)):
                party_name = str(row[1]).strip() if row[1] else ''
                vote_count = safe_int(row[2])
                vote_rate = float(row[3]) if row[3] else 0
                if party_name:
                    party_votes[party_name] = {'votes': vote_count, 'voteRate': round(vote_rate, 2)}

        # Get elected seats from party tables
        party_seats = {}
        table_nums = HIREI_TABLE_MAP[block_name]

        for tnum in table_nums:
            ws = wb[f'Table {tnum}']
            rows = list(ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), values_only=False))
            if len(rows) < 3:
                continue

            row1 = [(c.value, c.column) for c in rows[0]]
            row3 = [(c.value, c.column) for c in rows[2]]

            for val, col in row1:
                if not val:
                    continue
                for p in KNOWN_PARTIES:
                    if p in str(val):
                        seats = 0
                        for v3, c3 in row3:
                            if v3 and c3 >= col and c3 < col + 10:
                                m = re.search(r'(\d+)\s*人', str(v3))
                                if m:
                                    seats = int(m.group(1))
                                    break
                        if p not in party_seats:
                            party_seats[p] = seats
                        elif seats > 0:
                            party_seats[p] = seats
                        break

        # Build hirei parties
        hirei_parties = []
        for party_name, data in party_votes.items():
            seats = party_seats.get(party_name, 0)
            hirei_parties.append({
                'party': party_name,
                'block': block_name,
                'seats': seats,
                'votes': data['votes'],
                'voteRate': data['voteRate'],
                'candidates': [],
            })

        hirei_parties.sort(key=lambda x: (-x['seats'], -x['votes']))

        blocks.append({
            'name': block_name,
            'totalSeats': BLOCK_SEATS.get(block_name, 0),
            'totalVotes': total_votes,
            'parties': hirei_parties,
        })

    return blocks


def main():
    print('Loading Excel file...')
    wb = openpyxl.load_workbook(EXCEL_FILE)

    print('Extracting 小選挙区 data...')
    prefectures = extract_shou_data(wb)

    print('Extracting 比例代表 data...')
    hirei_blocks = extract_hirei_data(wb)

    total_shou_seats = sum(p['totalDistricts'] for p in prefectures)
    total_hirei_seats = sum(b['totalSeats'] for b in hirei_blocks)

    data = {
        'year': 2024,
        'electionDate': '2024-10-27',
        'hirei': {
            'totalSeats': total_hirei_seats,
            'blocks': hirei_blocks,
        },
        'shou': {
            'totalSeats': total_shou_seats,
            'prefectures': prefectures,
            'districts': [],
        },
    }

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f'\nOutput: {OUTPUT}')
    print(f'小選挙区: {total_shou_seats} seats, {len(prefectures)} prefectures')
    print(f'比例代表: {total_hirei_seats} seats, {len(hirei_blocks)} blocks')

    # Validation
    print('\n--- Validation ---')
    shou_ok = True
    for p in prefectures:
        total_elected = sum(r['seats'] for r in p['partyResults'])
        if total_elected != p['totalDistricts']:
            print(f"  WARN {p['prefecture']}: elected={total_elected} != districts={p['totalDistricts']}")
            shou_ok = False
    if shou_ok:
        print('  小選挙区: All prefectures OK')

    hirei_ok = True
    for b in hirei_blocks:
        total_elected = sum(p['seats'] for p in b['parties'])
        if total_elected != b['totalSeats']:
            print(f"  WARN {b['name']}: elected={total_elected} != seats={b['totalSeats']}")
            hirei_ok = False
    if hirei_ok:
        print('  比例代表: All blocks OK')

    # Summary
    print('\n--- 小選挙区 当選者数 ---')
    party_totals = {}
    for p in prefectures:
        for r in p['partyResults']:
            party_totals[r['party']] = party_totals.get(r['party'], 0) + r['seats']
    for party, seats in sorted(party_totals.items(), key=lambda x: -x[1]):
        if seats > 0:
            print(f'  {party}: {seats}')

    print('\n--- 比例代表 当選者数 ---')
    hirei_totals = {}
    for b in hirei_blocks:
        for p in b['parties']:
            hirei_totals[p['party']] = hirei_totals.get(p['party'], 0) + p['seats']
    for party, seats in sorted(hirei_totals.items(), key=lambda x: -x[1]):
        if seats > 0:
            print(f'  {party}: {seats}')


if __name__ == '__main__':
    main()
