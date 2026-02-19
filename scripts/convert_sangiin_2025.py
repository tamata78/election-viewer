#!/usr/bin/env python3
"""
2025年参議院選挙 全国データ変換スクリプト
Excel → NationalElectionData JSON

参議院選挙の特徴:
- 選挙区は合区あり (鳥取・島根、徳島・高知)
- 比例代表は全国一本 (ブロックなし)
"""
import json
import re
import unicodedata
import openpyxl

EXCEL_FILE = '/Users/tamata78/Downloads/2025_参議委員選挙_小選挙区_比例区.xlsx'
OUTPUT = '/Users/tamata78/work/election-viewer/public/data/elections/sangiin_2025.json'

# 45選挙区 (合区: 鳥取・島根、徳島・高知) — Tables 21-23 の行順
SANGIIN_PREFECTURES = [
    '北海道', '青森県', '岩手県', '宮城県', '秋田県', '山形県', '福島県',
    '茨城県', '栃木県', '群馬県', '埼玉県', '千葉県', '東京都', '神奈川県',
    '新潟県', '富山県', '石川県', '福井県', '山梨県', '長野県',
    '岐阜県', '静岡県', '愛知県', '三重県',
    '滋賀県', '京都府', '大阪府', '兵庫県', '奈良県', '和歌山県',
    '鳥取県・島根県', '岡山県', '広島県', '山口県',
    '徳島県・高知県', '香川県', '愛媛県',
    '福岡県', '佐賀県', '長崎県', '熊本県', '大分県', '宮崎県', '鹿児島県', '沖縄県',
]

# 標準47都道府県 (Tables 28, 38 の行順)
STANDARD_47 = [
    '北海道', '青森県', '岩手県', '宮城県', '秋田県', '山形県', '福島県',
    '茨城県', '栃木県', '群馬県', '埼玉県', '千葉県', '東京都', '神奈川県',
    '新潟県', '富山県', '石川県', '福井県', '山梨県', '長野県',
    '岐阜県', '静岡県', '愛知県', '三重県',
    '滋賀県', '京都府', '大阪府', '兵庫県', '奈良県', '和歌山県',
    '鳥取県', '島根県', '岡山県', '広島県', '山口県',
    '徳島県', '香川県', '愛媛県', '高知県',
    '福岡県', '佐賀県', '長崎県', '熊本県', '大分県', '宮崎県', '鹿児島県', '沖縄県',
]

# 合区マッピング
MERGE_MAP = {
    '鳥取県': '鳥取県・島根県',
    '島根県': '鳥取県・島根県',
    '徳島県': '徳島県・高知県',
    '高知県': '徳島県・高知県',
}


def normalize_party(name):
    """全角→半角変換等"""
    return unicodedata.normalize('NFKC', name).strip()


def safe_int(val):
    """小さな整数値 (当選者数など) を安全にパース"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace(',', '').replace(' ', '').replace('\u3000', '')
    m = re.search(r'\d+', s)
    return int(m.group(0)) if m else 0


def parse_count(val):
    """結合セル '2        1        3' から最後の数値 (計) を取得"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    nums = re.findall(r'\d+', s)
    if nums:
        return int(nums[-1])
    return 0


def parse_teisuu(val):
    """定数 '6(1)' → 7 (本来の定数+補欠の合計)"""
    if val is None:
        return 0
    s = str(val).strip()
    nums = re.findall(r'\d+', s)
    return sum(int(n) for n in nums) if nums else 0


def safe_votes(val):
    """得票数をパース。

    このExcelでは数値セルのピリオドが千区切りとして使われている:
    879.676 (float) → 879,676票
    文字列は通常のカンマ区切り: '14,470,016.925' → 14470017
    """
    if val is None:
        return 0
    if isinstance(val, str):
        s = val.replace(',', '').replace(' ', '').replace('\u3000', '')
        try:
            return int(round(float(s)))
        except (ValueError, TypeError):
            return 0
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        # ピリオドが千区切りのケース: 879.676 → 879676
        rounded = round(val, 3)
        s = f'{rounded:.3f}'
        return int(s.replace('.', ''))
    return 0


def get_elected(row, col_start, col_end):
    """カラム範囲から当選者数 (計) を取得。

    結合セル形式 (1カラムに '新 現 計') と
    分離カラム形式の両方に対応。後方から探して最初のnon-None値をparse。
    """
    for col in range(col_end, col_start - 1, -1):
        if col < len(row) and row[col] is not None:
            return parse_count(row[col])
    return 0


def extract_senkyoku_data(wb):
    """選挙区データ抽出"""
    # Tables 21-23: 当選者数 (45行, rows 3-47)
    rows21 = list(wb['Table 21'].iter_rows(min_row=3, max_row=47, values_only=True))
    rows22 = list(wb['Table 22'].iter_rows(min_row=3, max_row=47, values_only=True))
    rows23 = list(wb['Table 23'].iter_rows(min_row=3, max_row=47, values_only=True))

    # Table 28: 得票数 主要4党 (47行, rows 3-49)
    # col 7=自民計, col 10=立憲計, col 13=維新計, col 16=公明計
    rows28 = list(wb['Table 28'].iter_rows(min_row=3, max_row=49, values_only=True))

    # Table 38: 有効投票数 (47行, rows 3-49)
    # col 6=有効投票数
    rows38 = list(wb['Table 38'].iter_rows(min_row=3, max_row=49, values_only=True))

    # --- 47都道府県の得票データ (Table 28) ---
    pref_votes_47 = {}
    for i, pref in enumerate(STANDARD_47):
        r = rows28[i]
        pref_votes_47[pref] = {
            '自由民主党': safe_votes(r[7]),
            '立憲民主党': safe_votes(r[10]),
            '日本維新の会': safe_votes(r[13]),
            '公明党': safe_votes(r[16]),
        }

    # --- 47都道府県の有効投票数 (Table 38) ---
    pref_valid_47 = {}
    for i, pref in enumerate(STANDARD_47):
        r = rows38[i]
        pref_valid_47[pref] = safe_votes(r[6])

    # --- 合区マージ → 45選挙区 ---
    pref_votes_45 = {}
    pref_valid_45 = {}
    for pref in STANDARD_47:
        merged = MERGE_MAP.get(pref)
        target = merged if merged else pref
        if target not in pref_votes_45:
            pref_votes_45[target] = {}
            pref_valid_45[target] = 0
        for party, v in pref_votes_47[pref].items():
            pref_votes_45[target][party] = pref_votes_45[target].get(party, 0) + v
        pref_valid_45[target] += pref_valid_47[pref]

    # --- 45選挙区データ構築 ---
    prefectures = []
    for i, pref_name in enumerate(SANGIIN_PREFECTURES):
        r21 = rows21[i]
        r22 = rows22[i]
        r23 = rows23[i]

        teisuu = parse_teisuu(r21[5])

        # 当選者数
        elected = {
            '自由民主党': get_elected(r21, 6, 10),
            '立憲民主党': get_elected(r21, 11, 15),
            '日本維新の会': parse_count(r21[16]) if len(r21) > 16 else 0,
            '公明党': parse_count(r21[17]) if len(r21) > 17 else 0,
            '国民民主党': get_elected(r21, 18, 23),
            '日本共産党': parse_count(r21[24]) if len(r21) > 24 else 0,
            'れいわ新選組': get_elected(r21, 25, 26),
            '参政党': parse_count(r22[6]) if len(r22) > 6 else 0,
            '日本保守党': parse_count(r22[7]) if len(r22) > 7 else 0,
            '社会民主党': parse_count(r22[8]) if len(r22) > 8 else 0,
            'みんなでつくる党': parse_count(r22[9]) if len(r22) > 9 else 0,
            '無所属連合': parse_count(r22[10]) if len(r22) > 10 else 0,
            'チームみらい': parse_count(r22[11]) if len(r22) > 11 else 0,
            '日本誠真会': parse_count(r22[12]) if len(r22) > 12 else 0,
            '日本改革党': parse_count(r23[6]) if len(r23) > 6 else 0,
            '再生の道': parse_count(r23[7]) if len(r23) > 7 else 0,
            'NHK党': parse_count(r23[8]) if len(r23) > 8 else 0,
            '諸派': parse_count(r23[9]) if len(r23) > 9 else 0,
            '無所属': parse_count(r23[10]) if len(r23) > 10 else 0,
        }

        votes = pref_votes_45.get(pref_name, {})
        total_valid = pref_valid_45.get(pref_name, 0)
        if total_valid == 0:
            total_valid = sum(votes.values()) or 1

        # partyResults 構築 (議席 or 得票 がある政党のみ)
        party_results = []
        for party, e in elected.items():
            v = votes.get(party, 0)
            if e > 0 or v > 0:
                rate = (v / total_valid * 100) if total_valid > 0 else 0
                party_results.append({
                    'party': party,
                    'seats': e,
                    'totalVotes': v,
                    'voteRate': round(rate, 2),
                })

        party_results.sort(key=lambda x: (-x['seats'], -x['totalVotes']))

        prefectures.append({
            'prefecture': pref_name,
            'totalDistricts': teisuu,
            'partyResults': party_results,
        })

    return prefectures


def extract_hirei_data(wb):
    """比例代表データ抽出 (全国一本)"""
    # Table 40: 政党別得票 (ランキング)
    party_data = {}

    # 左側 (rows 3-16, rank 1-14)
    for row in wb['Table 40'].iter_rows(min_row=3, max_row=16, values_only=True):
        if isinstance(row[0], (int, float)) and row[1]:
            name = normalize_party(str(row[1]))
            votes = safe_votes(row[2])
            rate = float(str(row[3]).strip()) if row[3] else 0
            party_data[name] = {'votes': votes, 'voteRate': round(rate, 2), 'seats': 0}

    # 右側 (rows 3-4, rank 15-16)
    for row in wb['Table 40'].iter_rows(min_row=3, max_row=4, values_only=True):
        if isinstance(row[4], (int, float)) and row[5]:
            name = normalize_party(str(row[5]))
            if name == '得票総数':
                continue
            votes = safe_votes(row[6])
            rate = float(str(row[7]).strip()) if row[7] else 0
            party_data[name] = {'votes': votes, 'voteRate': round(rate, 2), 'seats': 0}

    # 得票総数 (Table 40 Row 5, col 6)
    r5 = list(wb['Table 40'].iter_rows(min_row=5, max_row=5, values_only=True))[0]
    total_votes = safe_votes(r5[6])

    # 当選人数 (Tables 41-44 Row 6)
    r6_41 = list(wb['Table 41'].iter_rows(min_row=6, max_row=6, values_only=True))[0]
    r6_42 = list(wb['Table 42'].iter_rows(min_row=6, max_row=6, values_only=True))[0]
    r6_43 = list(wb['Table 43'].iter_rows(min_row=6, max_row=6, values_only=True))[0]
    r6_44 = list(wb['Table 44'].iter_rows(min_row=6, max_row=6, values_only=True))[0]

    hirei_seats = {
        '自由民主党': safe_int(r6_41[3]),
        '立憲民主党': safe_int(r6_41[9]),
        '日本維新の会': safe_int(r6_41[15]),
        '公明党': safe_int(r6_41[22]),
        '国民民主党': safe_int(r6_42[3]),
        '日本共産党': safe_int(r6_42[10]),
        'れいわ新選組': safe_int(r6_42[16]),
        '参政党': safe_int(r6_42[22]),
        '日本保守党': safe_int(r6_43[3]),
        '社会民主党': safe_int(r6_43[9]),
        '無所属連合': safe_int(r6_43[15]),
        'チームみらい': safe_int(r6_43[21]),
        '日本誠真会': safe_int(r6_44[3]),
        '日本改革党': safe_int(r6_44[9]),
        '再生の道': safe_int(r6_44[15]),
        'NHK党': safe_int(r6_44[21]),
    }

    # 当選人数を party_data にマージ
    for name, seats in hirei_seats.items():
        if name in party_data:
            party_data[name]['seats'] = seats
        else:
            print(f'  WARN: {name} not found in Table 40 votes')

    # parties リスト構築
    parties = []
    for name, data in party_data.items():
        parties.append({
            'party': name,
            'block': '全国',
            'seats': data['seats'],
            'votes': data['votes'],
            'voteRate': data['voteRate'],
            'candidates': [],
        })

    parties.sort(key=lambda x: (-x['seats'], -x['votes']))
    total_seats = sum(p['seats'] for p in parties)

    block = {
        'name': '全国',
        'totalSeats': total_seats,
        'totalVotes': total_votes,
        'parties': parties,
    }

    return [block], total_seats


def main():
    print('Loading Excel file...')
    wb = openpyxl.load_workbook(EXCEL_FILE)

    print('Extracting 選挙区 data...')
    prefectures = extract_senkyoku_data(wb)

    print('Extracting 比例代表 data...')
    hirei_blocks, total_hirei_seats = extract_hirei_data(wb)

    total_senkyoku_seats = sum(p['totalDistricts'] for p in prefectures)

    data = {
        'year': 2025,
        'electionDate': '2025-07-20',
        'hirei': {
            'totalSeats': total_hirei_seats,
            'blocks': hirei_blocks,
        },
        'shou': {
            'totalSeats': total_senkyoku_seats,
            'prefectures': prefectures,
            'districts': [],
        },
    }

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f'\nOutput: {OUTPUT}')
    print(f'選挙区: {total_senkyoku_seats} seats, {len(prefectures)} prefectures')
    print(f'比例代表: {total_hirei_seats} seats, {len(hirei_blocks)} blocks')

    # --- Validation ---
    print('\n--- Validation ---')

    # 選挙区: 当選者数 == 定数
    senkyoku_ok = True
    for p in prefectures:
        total_elected = sum(r['seats'] for r in p['partyResults'])
        if total_elected != p['totalDistricts']:
            print(f"  WARN {p['prefecture']}: elected={total_elected} != teisuu={p['totalDistricts']}")
            senkyoku_ok = False
    if senkyoku_ok:
        print('  選挙区: All prefectures OK')

    # 比例代表: 当選者数 == 50
    hirei_ok = True
    for b in hirei_blocks:
        total_elected = sum(p['seats'] for p in b['parties'])
        if total_elected != b['totalSeats']:
            print(f"  WARN {b['name']}: elected={total_elected} != seats={b['totalSeats']}")
            hirei_ok = False
    if hirei_ok:
        print('  比例代表: All blocks OK')

    # --- Summary ---
    print('\n--- 選挙区 当選者数 ---')
    party_totals = {}
    for p in prefectures:
        for r in p['partyResults']:
            party_totals[r['party']] = party_totals.get(r['party'], 0) + r['seats']
    for party, seats in sorted(party_totals.items(), key=lambda x: -x[1]):
        if seats > 0:
            print(f'  {party}: {seats}')

    print('\n--- 比例代表 当選者数 ---')
    for b in hirei_blocks:
        for p in b['parties']:
            if p['seats'] > 0:
                print(f"  {p['party']}: {p['seats']}")


if __name__ == '__main__':
    main()
